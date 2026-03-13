from odoo import models, fields
import base64
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from odoo.exceptions import UserError
from datetime import datetime


class AppointmentReportWizard(models.TransientModel):
    _name = 'appointment.report.wizard'
    _description = 'Appointment Report Wizard'


    doctor_ids = fields.Many2many(
        'res.partner',
        'appointment_report_doctor_rel',
        'wizard_id',
        'doctor_id',
        string="Doctors",
        domain="[('role_ids.name','=','Doctor')]"
    )

    patient_ids = fields.Many2many(
        'res.partner',
        'appointment_report_patient_rel',
        'wizard_id',
        'patient_id',
        string="Patients",
        domain="[('role_ids.name','=','Patient')]"
    )

    date_from = fields.Date(string="Date From")
    date_to = fields.Date(string="Date To")

    status_ids = fields.Many2many(
        'hospital.appointment.status',
        'appointment_report_status_rel',
        'wizard_id',
        'status_id',
        string="Status"
    )


    def action_print_report(self):

        domain = []

        # Doctor login restriction
        if self.env.user.has_group('hospital_management.group_hospital_doctor'):
            domain.append(('doctor_id.user_id', '=', self.env.user.id))

        elif self.doctor_ids:
            domain.append(('doctor_id', 'in', self.doctor_ids.ids))

        if self.patient_ids:
            domain.append(('patient_id', 'in', self.patient_ids.ids))

        if self.status_ids:
            domain.append(('state', 'in', self.status_ids.mapped('code')))

        if self.date_from:
            domain.append(('start_date', '>=', self.date_from))

        if self.date_to:
            domain.append(('start_date', '<=', self.date_to))


        appointments = self.env['hospital.appointment'].sudo().search(domain)

        # IMPORTANT VALIDATION
        if not appointments:
            raise UserError("No Appointment Found!")


        data = {
            'ids': appointments.ids,
            'model': 'hospital.appointment',
        }

        return self.env.ref(
            'hospital_management.action_appointment_summary_report'
        ).report_action(appointments, data=data)




    def action_export_excel(self):

        domain = []

        # Doctor login restriction
        if self.env.user.has_group('hospital_management.group_hospital_doctor'):
            domain.append(('doctor_id.user_id', '=', self.env.user.id))

        elif self.doctor_ids:
            domain.append(('doctor_id', 'in', self.doctor_ids.ids))

        if self.patient_ids:
            domain.append(('patient_id', 'in', self.patient_ids.ids))

        if self.status_ids:
            domain.append(('state', 'in', self.status_ids.mapped('code')))

        if self.date_from:
            domain.append(('start_date', '>=', self.date_from))

        if self.date_to:
            domain.append(('start_date', '<=', self.date_to))

        appointments = self.env['hospital.appointment'].search(domain)

        if not appointments:
            raise UserError("No Appointment Found!")

        wb = Workbook()
        sheet = wb.active
        sheet.title = "Appointments"

        # Column width
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 30
        sheet.column_dimensions['D'].width = 25
        sheet.column_dimensions['E'].width = 16
        sheet.column_dimensions['F'].width = 15

        # Styles
        title_font = Font(size=16, bold=True, color="2A7F62")
        header_font = Font(size=13, color="FFFFFF")
        data_font = Font(size=12)

        header_fill = PatternFill(start_color="2A7F62", end_color="2A7F62", fill_type="solid")

        center_align = Alignment(vertical="center")
        right_align = Alignment(horizontal="right")

        border = Border(bottom=Side(style='thin', color="DDDDDD"))


        # ===== Title =====
        sheet.merge_cells('A1:C2')

        title = sheet['A1']
        title.value = "Appointment Summary Report"
        title.font = title_font
        title.alignment = Alignment(horizontal="left", vertical="center")


        bold = InlineFont(b=True)
        normal = InlineFont()

        # Generated On
        sheet.merge_cells('E1:F1')
        gen = sheet['E1']
        gen.value = CellRichText(
            TextBlock(bold, "Generated On : "),
            TextBlock(
                normal,
                fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()
                ).strftime("%d-%m-%Y %I:%M %p")
            )
        )       
        gen.alignment = Alignment(horizontal="left")

        # Total Appointments
        sheet.merge_cells('E2:F2')
        total_cell = sheet['E2']
        total_cell.value = CellRichText(
            TextBlock(bold, "Total Appointments : "),
            TextBlock(normal, str(len(appointments)))
        )
        total_cell.alignment = Alignment(horizontal="left")



        # ===== Header =====
        headers = ["ID", "Doctor", "Patient", "Date", "Fees", "Status"]

        header_row = 4

        for col, header in enumerate(headers, 1):

            cell = sheet.cell(row=header_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Row height (header bigger)
        sheet.row_dimensions[header_row].height = 25

        # 👉 Freeze top rows
        sheet.freeze_panes = "A5"

        # ===== Data =====
        row = header_row + 1

        for rec in appointments:

            c1 = sheet.cell(row=row, column=1, value=rec.name)
            
            doctor_name = f"[{rec.doctor_id.hospital_code}] {rec.doctor_id.name}"
            c2 = sheet.cell(row=row, column=2, value=doctor_name)

            patient_name = f"[{rec.patient_id.hospital_code}] {rec.patient_id.name}"
            c3 = sheet.cell(row=row, column=3, value=patient_name)

            if rec.start_date:
                user_time = fields.Datetime.context_timestamp(self, rec.start_date)

                c4 = sheet.cell(
                    row=row,
                    column=4,
                    value=user_time.strftime("%m/%d/%Y %I:%M:%S %p")
                )
            else:
                c4 = sheet.cell(row=row, column=4, value="")

            c5 = sheet.cell(row=row, column=5, value=rec.fees)
            # Status label
            state_label = dict(rec._fields['state'].selection).get(rec.state)

            c6 = sheet.cell(row=row, column=6, value=state_label)

            for c in [c1, c2, c3, c4, c5, c6]:
                c.font = data_font
                c.border = border
                c.alignment = center_align
                
            c5.font = data_font
            c5.border = border
            c5.alignment = Alignment(vertical="center")
            c5.number_format = '₹#,##0.00'

            sheet.row_dimensions[row].height = 24

            row += 1


        # Save Excel
        file_data = BytesIO()
        wb.save(file_data)
        file_data.seek(0)

        attachment = self.env['ir.attachment'].create({
            'name': 'Appointment_Report.xlsx',
            'type': 'binary',
            'datas': base64.b64encode(file_data.read()),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        })

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'self',
        }